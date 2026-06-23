"""
Excel 模板处理器 - 通用脚本
基于 NAT6602~NAT6607 生成经验提炼

功能：
1. 从模板复制（保留原始结构）
2. 读取源数据（Excel）
3. 提取图片到临时文件
4. 捕获模板格式（合并单元格、行高、列宽、样式）
5. 生成新文件（填充数据 + 嵌入图片）
6. 避免文件占用问题
"""

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
from openpyxl.drawing.spreadsheet_drawing import TwoCellAnchor, AnchorMarker
from copy import copy
import io
import os
import shutil
import tempfile
from PIL import Image as PILImage


def read_product_list(product_list_path, min_row=3, max_row=30, max_col=7):
    """
    读取产品清单
    返回：{序号: {'partno': 零件号, 'defects_raw': 缺陷描述, ...}}
    """
    print(f"读取产品清单: {product_list_path}")
    wb = openpyxl.load_workbook(product_list_path)
    ws = wb.active
    products = {}
    
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, max_col=max_col, values_only=True):
        seq = row[0]
        if seq is None:
            continue
        try:
            seq = int(seq)
        except Exception:
            continue
        
        products[seq] = {
            'seq': seq,
            'partno': str(row[1]).strip() if len(row) > 1 and row[1] else "",
            'defects_raw': str(row[4]).strip() if len(row) > 4 and row[4] else "",
        }
    
    wb.close()
    print(f"  读取到 {len(products)} 个产品")
    return products


def extract_images_from_excel(product_list_path, output_dir=None, max_seq=28):
    """
    从 Excel 中提取嵌入的图片
    返回：{序号: [图片路径1, 图片路径2, ...]}
    """
    if output_dir is None:
        output_dir = tempfile.mkdtemp(prefix="excel_img_")
    
    print(f"提取图片到: {output_dir}")
    img_by_seq = {}
    
    wb = openpyxl.load_workbook(product_list_path)
    ws = wb.active
    
    for img in ws._images:
        anc = img.anchor
        if hasattr(anc, '_from'):
            img_row = anc._from.row + 1
            product_seq = img_row - 2  # 根据模板调整
            
            if 1 <= product_seq <= max_seq:
                try:
                    img_data = img._data()
                    if product_seq not in img_by_seq:
                        img_by_seq[product_seq] = []
                    
                    tmp_path = os.path.join(output_dir, f"img_{product_seq}_{len(img_by_seq[product_seq])}.png")
                    with open(tmp_path, 'wb') as f:
                        f.write(img_data)
                    
                    img_by_seq[product_seq].append(tmp_path)
                except Exception as e:
                    print(f"  WARN: 提取序号 {product_seq} 图片失败: {e}")
    
    wb.close()
    print(f"  提取完成，共 {sum(len(v) for v in img_by_seq.values())} 张")
    return img_by_seq


def capture_template_format(template_path, product_start_row=39, product_end_row=45):
    """
    捕获模板格式
    返回：{
        'merges': [合并单元格信息],
        'row_heights': {行偏移: 行高},
        'col_widths': {列号: 列宽},
        'styles': {(行偏移, 列偏移): 样式},
        'header_heights': {行号: 行高}
    }
    """
    print(f"捕获模板格式: {template_path}")
    wb = openpyxl.load_workbook(template_path)
    ws = wb.active
    
    template = {
        'merges': [],
        'row_heights': {},
        'col_widths': {},
        'styles': {},
        'header_heights': {}
    }
    
    # 捕获产品区域合并单元格
    for mc in ws.merged_cells.ranges:
        if product_start_row <= mc.min_row <= product_end_row:
            template['merges'].append({
                'min_row_offset': mc.min_row - product_start_row,
                'max_row_offset': mc.max_row - product_start_row,
                'min_col': mc.min_col,
                'max_col': mc.max_col,
            })
    
    # 捕获行高
    for r in range(product_start_row, product_end_row + 1):
        if ws.row_dimensions[r].height:
            template['row_heights'][r - product_start_row] = ws.row_dimensions[r].height
    
    # 捕获样式
    for r in range(product_start_row, product_end_row + 1):
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(row=r, column=c)
            template['styles'][(r - product_start_row, c - 1)] = {
                'font': copy(cell.font) if cell.font else None,
                'fill': copy(cell.fill) if cell.fill else None,
                'alignment': copy(cell.alignment) if cell.alignment else None,
                'border': copy(cell.border) if cell.border else None,
                'number_format': cell.number_format,
            }
    
    # 捕获头部格式
    for r in range(1, product_start_row):
        if ws.row_dimensions[r].height:
            template['header_heights'][r] = ws.row_dimensions[r].height
    
    # 捕获列宽
    for c in range(1, ws.max_column + 1):
        cl = get_column_letter(c)
        if cl in ws.column_dimensions:
            template['col_widths'][c] = ws.column_dimensions[cl].width
    
    wb.close()
    print(f"  捕获完成: {len(template['merges'])} 个合并单元格, {len(template['styles'])} 个样式")
    return template


def add_image_to_sheet(ws, img_path, slot_left, slot_top, slot_right, slot_bottom):
    """
    添加图片到工作表（使用 TwoCellAnchor）
    
    参数：
    - ws: 工作表对象
    - img_path: 图片路径
    - slot_left, slot_top: 锚定区域左上角（0-indexed）
    - slot_right, slot_bottom: 锚定区域右下角（0-indexed）
    """
    if not os.path.exists(img_path):
        print(f"  WARN: 图片不存在: {img_path}")
        return False
    
    try:
        # 读取并调整图片大小
        pil_img = PILImage.open(img_path)
        
        # 调整大小
        target_width = 200
        target_height = 130
        
        img_width, img_height = pil_img.size
        scale = min(target_width / img_width, target_height / img_height)
        new_width = int(img_width * scale)
        new_height = int(img_height * scale)
        
        pil_img = pil_img.resize((new_width, new_height), PILImage.LANCZOS)
        
        # 保存到临时文件
        temp_fd, temp_path = tempfile.mkstemp(suffix=".png")
        os.close(temp_fd)
        pil_img.save(temp_path)
        
        # 创建 XLImage
        img = XLImage(temp_path)
        
        # 设置锚点
        img.anchor = TwoCellAnchor(
            _from=AnchorMarker(row=slot_top, col=slot_left),
            to=AnchorMarker(row=slot_bottom, col=slot_right),
        )
        
        # 添加到工作表
        ws.add_image(img)
        
        print(f"  已添加图片: {os.path.basename(img_path)}")
        return True
        
    except Exception as e:
        print(f"  ERROR: 添加图片失败 {img_path}: {e}")
        return False


def generate_files_from_template(template_path, product_list_path, output_dir, 
                               groups, file_names, rows_per_product=7,
                               product_start_row=39):
    """
    从模板批量生成文件（主函数）
    
    参数：
    - template_path: 模板路径
    - product_list_path: 产品清单路径
    - output_dir: 输出目录
    - groups: 分组列表 [[序号1, 序号2, ...], ...]
    - file_names: 文件名列表 ["NAT6602", "NAT6603", ...]
    - rows_per_product: 每个产品占用的行数
    - product_start_row: 产品区域起始行
    """
    print("="*50)
    print("开始生成文件...")
    print("="*50)
    
    # STEP 1: 读取产品清单
    products = read_product_list(product_list_path)
    
    # STEP 2: 提取图片
    img_by_seq = extract_images_from_excel(product_list_path)
    
    # STEP 3: 捕获模板格式
    template = capture_template_format(template_path, product_start_row)
    
    # STEP 4: 生成文件
    os.makedirs(output_dir, exist_ok=True)
    
    for group_idx, seqs in enumerate(groups):
        if group_idx >= len(file_names):
            break
        
        fname = file_names[group_idx]
        num_products = len(seqs)
        
        print(f"\n--- {fname}.xlsx ({group_idx+1}/{len(groups)} 组, {num_products} 个产品) ---")
        
        # 输出路径
        output_path = os.path.join(output_dir, f"{fname}.xlsx")
        
        # 如果文件已存在，删除
        if os.path.exists(output_path):
            try:
                os.remove(output_path)
                print(f"  已删除旧文件: {output_path}")
            except Exception as e:
                print(f"  WARN: 无法删除 {output_path}: {e}")
                continue
        
        # 复制模板
        shutil.copy2(template_path, output_path)
        print(f"  已复制模板到: {output_path}")
        
        # 加载副本
        wb = openpyxl.load_workbook(output_path)
        ws = wb.active
        
        # 清除现有图片
        if hasattr(ws, '_images'):
            old_count = len(ws._images)
            ws._images = []
            print(f"  已清除 {old_count} 个旧图片")
        
        # 删除旧数据行
        old_max = ws.max_row
        if old_max >= product_start_row:
            ws.delete_rows(product_start_row, old_max - product_start_row + 1)
        
        # 插入新行
        total_new_rows = num_products * rows_per_product
        if total_new_rows > 0:
            ws.insert_rows(product_start_row, total_new_rows)
        
        # 还原头部格式
        for r, h in template['header_heights'].items():
            if r < product_start_row:
                ws.row_dimensions[r].height = h
        for c, w in template['col_widths'].items():
            ws.column_dimensions[get_column_letter(c)].width = w
        
        # 按产品填充
        for pi, seq in enumerate(seqs):
            prod = products.get(seq, {})
            if not prod:
                print(f"  WARN: 序号 {seq} 未找到，跳过")
                continue
            
            start_row = product_start_row + pi * rows_per_product
            
            # 恢复合并单元格
            for tm in template['merges']:
                m_min_row = start_row + tm['min_row_offset']
                m_max_row = start_row + tm['max_row_offset']
                m_min_col = tm['min_col']
                m_max_col = tm['max_col']
                m_sc = get_column_letter(m_min_col)
                m_ec = get_column_letter(m_max_col)
                ws.merge_cells(f"{m_sc}{m_min_row}:{m_ec}{m_max_row}")
            
            # 恢复行高
            for r_off in range(rows_per_product):
                rn = start_row + r_off
                if r_off in template['row_heights']:
                    ws.row_dimensions[rn].height = template['row_heights'][r_off]
            
            # 恢复样式
            for r_off in range(rows_per_product):
                rn = start_row + r_off
                for c_off in range(ws.max_column):
                    cn = c_off + 1
                    cell = ws.cell(row=rn, column=cn)
                    sk = (r_off, c_off)
                    if sk in template['styles']:
                        s = template['styles'][sk]
                        if s['font']:
                            cell.font = copy(s['font'])
                        if s['fill']:
                            cell.fill = copy(s['fill'])
                        if s['alignment']:
                            cell.alignment = copy(s['alignment'])
                        if s['border']:
                            cell.border = copy(s['border'])
                        cell.number_format = s['number_format']
            
            # 填入数据
            ws.cell(row=start_row, column=1, value=prod.get('partno', ''))
            
            # 添加图片
            if seq in img_by_seq and img_by_seq[seq]:
                print(f"  产品 序号{seq}: 找到 {len(img_by_seq[seq])} 张图片")
                
                # 计算图片槽位（H:M 区域，每行最多3张）
                num_imgs = len(img_by_seq[seq])
                imgs_per_row = min(num_imgs, 3)
                
                for idx, img_path in enumerate(img_by_seq[seq]):
                    row_idx = idx // imgs_per_row
                    col_idx = idx % imgs_per_row
                    
                    # H=8, M=13 (1-indexed) -> 7, 12 (0-indexed)
                    region_left = 7
                    region_right = 12
                    region_top = start_row - 1
                    region_bottom = start_row + rows_per_product - 2
                    
                    total_cols = region_right - region_left + 1
                    slot_width = total_cols // imgs_per_row
                    
                    slot_left = region_left + col_idx * slot_width
                    slot_right = slot_left + slot_width - 1
                    
                    total_rows = region_bottom - region_top + 1
                    num_rows = (num_imgs + imgs_per_row - 1) // imgs_per_row
                    slot_height = total_rows // num_rows
                    
                    slot_top = region_top + row_idx * slot_height
                    slot_bottom = slot_top + slot_height - 1
                    
                    add_image_to_sheet(ws, img_path, slot_left, slot_top, slot_right, slot_bottom)
            else:
                print(f"  产品 序号{seq}: 无图片")
        
        # 保存
        try:
            wb.save(output_path)
            print(f"  已保存: {output_path}")
        except Exception as e:
            print(f"  ERROR: 保存失败: {e}")
            continue
    
    print("\n" + "="*50)
    print("生成完成！")
    print(f"输出目录: {output_dir}")
    print("="*50)


# ===========================================================
# 使用示例
# ===========================================================
if __name__ == "__main__":
    # 配置（根据实际情况修改）
    BASE_DIR = r"C:\Users\HUAWEI\Desktop\工作管理\日常进度\项目\重庆标准件工业有限责任公司重庆汽车标准件厂分公司"
    TEMPLATE = os.path.join(BASE_DIR, "NAT6602.xlsx")
    PRODUCT_LIST = os.path.join(BASE_DIR, "筛选机计划筛选的产品清单-2026.6.9.xlsx")
    OUTPUT_DIR = os.path.join(BASE_DIR, "_generated")
    
    ALL_GROUPS = [
        [1, 2, 3, 4],
        [5, 6, 7, 8, 9, 10, 11],
        [12, 13],
        [14, 15, 16],
        [17, 18, 19, 20, 21, 22, 23, 24, 25],
        [26, 27, 28],
    ]
    FILE_NAMES = ["NAT6602", "NAT6603", "NAT6604", "NAT6605", "NAT6606", "NAT6607"]
    
    # 生成文件
    generate_files_from_template(
        template_path=TEMPLATE,
        product_list_path=PRODUCT_LIST,
        output_dir=OUTPUT_DIR,
        groups=ALL_GROUPS,
        file_names=FILE_NAMES,
        rows_per_product=7,
        product_start_row=39
    )
